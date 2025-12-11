from pathlib import Path
import sqlite3
import pandas as pd
from openpyxl import load_workbook

EXCEL = Path(r"e:\GitHub-projects\analise-consumo-de-energia\Dados_abertos_Consumo_Mensal.xlsx")
OUT_DB = Path(r"e:\GitHub-projects\analise-consumo-de-energia\consumo_mensal.db")

if not EXCEL.exists():
    print(f"Excel file not found at {EXCEL}")
    raise SystemExit(1)

print(f"Importing sheets from: {EXCEL}")
wb = load_workbook(EXCEL, read_only=True, data_only=True)
sheets = wb.sheetnames
print("Sheets to import:", sheets)

conn = sqlite3.connect(OUT_DB)
for sheet in sheets:
    print(f"\nReading sheet: {sheet}")
    try:
        # Read one sheet at a time to limit memory usage
        df = pd.read_excel(EXCEL, sheet_name=sheet, engine='openpyxl')
    except Exception as e:
        print(f"  Failed to read sheet {sheet}: {e}")
        continue

    # Normalize column names
    cols = []
    for i, col in enumerate(df.columns):
        if col is None:
            name = f"col_{i}"
        else:
            name = str(col).strip()
            name = name.replace(' ', '_').replace('-', '_')
            # keep only alnum and underscore
            name = ''.join(c if (c.isalnum() or c == '_') else '_' for c in name)
            if name == '':
                name = f"col_{i}"
        cols.append(name)
    df.columns = cols

    # Normalize table name
    tname = sheet.strip().replace(' ', '_').replace('-', '_')
    tname = ''.join(c if (c.isalnum() or c == '_') else '_' for c in tname)
    if tname == '':
        tname = f"sheet_{sheets.index(sheet)}"

    print(f"  Writing to table: {tname} (rows: {len(df)})")
    try:
        df.to_sql(tname, conn, if_exists='replace', index=False)
        cur = conn.execute(f"SELECT COUNT(*) FROM '{tname}'")
        rc = cur.fetchone()[0]
        print(f"  Wrote {rc} rows to {tname}")
    except Exception as e:
        print(f"  Failed to write table {tname}: {e}")

conn.close()
print(f"Done. SQLite DB created at: {OUT_DB}")
