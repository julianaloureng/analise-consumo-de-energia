from openpyxl import load_workbook
from pathlib import Path

FILE = Path(r"e:\GitHub-projects\analise-consumo-de-energia\Dados_abertos_Consumo_Mensal.xlsx")

if not FILE.exists():
    print(f"Excel file not found at {FILE}")
    raise SystemExit(1)

print(f"Reading (fast mode): {FILE}")
wb = load_workbook(FILE, read_only=True, data_only=True)
print("Found sheets:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print("\n" + "-" * 60)
    print(f"Sheet: {name}")
    it = ws.iter_rows(values_only=True)
    # print up to first 5 non-empty rows
    count = 0
    for row in it:
        # skip entirely empty rows
        if row is None:
            continue
        if all(cell is None for cell in row):
            continue
        print(row)
        count += 1
        if count >= 5:
            break
